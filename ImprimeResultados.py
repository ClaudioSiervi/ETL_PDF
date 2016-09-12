# -*- coding: utf-8 -*-
"""
Created on Tue Aug 02 19:20:17 2016

@author: Claudio
"""
from openpyxl import load_workbook
#from Arquivo import ArquivoIPDO
from Utilitarios import Ferramentas
#from openpyxl import Workbook
#import string

class ImprimeArquivosTexto():
    
    
    # Salva o texto extraído do pdf em um arquivo texto
    def texto_em_txt(self, texto, nome_arquivo_saida):
    
        arquivo_texto = open(nome_arquivo_saida, "w")
        self.status_escrita = arquivo_texto.write(texto)
        arquivo_texto.close()
        
     # Salva o texto extraído do pdf em um arquivo texto
    def texto_em_html(self, texto, nome_arquivo_saida):
   
        arquivo_texto = open(nome_arquivo_saida, "w")
        self.status_escrita = arquivo_texto.write(texto)
        arquivo_texto.close()
        

    # Imprime o resumo do balanço energético    
    def balanco_energia_resumido_em_xlsx(self, geral, balanco_resumido):
        
        wb_ipdo = load_workbook('IPDO.xlsx') 
        ws_balanco_resumido = wb_ipdo['BalancoResumido']
        
        ferramenta = Ferramentas()
        
        [primeira_linha, ultima_linha] = ferramenta.linha_nao_vazia(ws_balanco_resumido)
        
        ultima_linha = ultima_linha + 1        
        
        indice = 'A'+str(ultima_linha) 
           
        ws_balanco_resumido[indice] = geral["data_arquivo"]
        
        num_elementos_pg = len(balanco_resumido["programada"]) # Número de elementos do resumo
        conta_valores = 0
        
        for conta_coluna_pg in xrange(1, num_elementos_pg):
            letra = ferramenta.retorna_letra_da_coluna(conta_coluna_pg + 1)
            indice = letra + str(ultima_linha)            
            ws_balanco_resumido[indice] = float(balanco_resumido["programada"][conta_valores])       
            conta_valores = conta_valores + 1
        
        num_elementos_vf = len(balanco_resumido["verificada"]) + conta_coluna_pg
        conta_valores = 0       
        
        for conta_coluna_vf in xrange(conta_coluna_pg, num_elementos_vf):
            letra = ferramenta.retorna_letra_da_coluna(conta_coluna_vf + 2)
            indice = letra + str(ultima_linha)            
#            print balanco_resumido["verificada"][conta_valores]
            ws_balanco_resumido[indice] = balanco_resumido["verificada"][conta_valores]
            conta_valores = conta_valores + 1                
        
        wb_ipdo.save('IPDO.xlsx')   # sobrescreve resultados

        
    # Imprime resultados na aba 'BalancoDetalhado' da planilha 'IPDO.xlsx'
    def balanco_energia_detalhado_em_xlsx(self, geral, balanco_detalhado):
        
        wb_ipdo = load_workbook('IPDO.xlsx') 
        ws_balanco_detalhado = wb_ipdo['BalancoDetalhado']
       
        ferramenta = Ferramentas()
        
        [primeira_linha, ultima_linha] = ferramenta.linha_nao_vazia(ws_balanco_detalhado)
        
        ultima_linha = ultima_linha + 1        
            
        indice = 'A'+str(ultima_linha)
        
        ws_balanco_detalhado[indice] = geral["data_arquivo"]
        
        fontes_relatorio = ["Hidro","Termo","Nuclear","Eólica","Solar","Total","Carga" ]
        subsistemas_relatorio = ["sudeste","sul","nordeste","norte"]
        
        coluna_energia_pg = 1 # primeira fonte programada
        coluna_carga_pg = 8        
        
        coluna_energia_vf = 8 # primeira fonte verificada
        coluna_carga_vf = 15
        
        estados = 2 # programada e verificada
        colunas_relatorio = estados * len(subsistemas_relatorio)*len(fontes_relatorio)
       
       # zera todas as células da ultima linha
        for coluna in xrange(2, (colunas_relatorio+2)):
            indice = ferramenta.retorna_letra_da_coluna(coluna) + str(ultima_linha)
            ws_balanco_detalhado[indice] = 0 
            
        for subsistema in subsistemas_relatorio:    

            for fonte in fontes_relatorio:
                coluna_energia_pg += 1
                coluna_energia_vf += 1
                
                indice_pg = ferramenta.retorna_letra_da_coluna(coluna_energia_pg) + str(ultima_linha)
                indice_vf = ferramenta.retorna_letra_da_coluna(coluna_energia_vf) + str(ultima_linha)
                
                indice_carga_pg = ferramenta.retorna_letra_da_coluna(coluna_carga_pg) + str(ultima_linha)
                indice_carga_vf = ferramenta.retorna_letra_da_coluna(coluna_carga_vf) + str(ultima_linha)

                for fonte_extraida in balanco_detalhado[subsistema]["energia"]:

                    if (fonte_extraida == fonte):
                        ws_balanco_detalhado[indice_pg] = float(balanco_detalhado[subsistema]["energia"][fonte]["programada"])
                        ws_balanco_detalhado[indice_vf] = float(balanco_detalhado[subsistema]["energia"][fonte]["verificada"])
                    
                    elif (fonte == "Carga"):                       
                        ws_balanco_detalhado[indice_carga_pg] = float(balanco_detalhado[subsistema]["carga"]["programada"])                        
                        ws_balanco_detalhado[indice_carga_vf] = float(balanco_detalhado[subsistema]["carga"]["verificada"])
                                                
                        continue
                        
            coluna_energia_pg += 7  # pula  fontes verificadas
            coluna_energia_vf += 7  # pula fontes programadas
            coluna_carga_pg += 14             
            coluna_carga_vf += 14 

        wb_ipdo.save('IPDO.xlsx')   # sobrescreve resultados
        
        
        
    def intercambio_em_xlsx(self, geral, balanco_detalhado):
        
        wb_ipdo = load_workbook('IPDO.xlsx') 
        ws_balanco_detalhado = wb_ipdo['Intercambios']
        
        ferramenta = Ferramentas()
        
        [primeira_linha, ultima_linha] = ferramenta.linha_nao_vazia(ws_balanco_detalhado)
        
        ultima_linha = ultima_linha + 1        
            
        indice = 'A'+str(ultima_linha)
        
        ws_balanco_detalhado[indice] = geral["data_arquivo"]
        
                                        # saida-chegada
        sentido_transferencia_energia = ['norte-imperatriz', 'imperatriz-nordeste', \
                                        'itaipu-sudeste','sudeste-imperatriz', 'sul-sudeste', \
                                        'internacional-sul']
        # referência excel
        coluna_energia_pg = 1 # primeira intercambio programada
        coluna_energia_vf = 7 # primeira intercambio verificada
        
        estados = 2 # programada e verificada
        colunas_relatorio = estados * len(sentido_transferencia_energia)
       
       # zera todas as células da ultima linha
        for coluna in xrange(2, (colunas_relatorio+2)):
            indice = ferramenta.retorna_letra_da_coluna(coluna) + str(ultima_linha)
            ws_balanco_detalhado[indice] = 0 
    
        for intercambio_energia in sentido_transferencia_energia:    
            coluna_energia_pg += 1
            coluna_energia_vf += 1
            
            indice_pg = ferramenta.retorna_letra_da_coluna(coluna_energia_pg) + str(ultima_linha)
            indice_vf = ferramenta.retorna_letra_da_coluna(coluna_energia_vf) + str(ultima_linha)
            
            print intercambio_energia
            
            for intercambio_extraido in balanco_detalhado['intercambio']:
                
                if (intercambio_extraido == intercambio_energia):
                    print balanco_detalhado['intercambio'][intercambio_extraido]
                    ws_balanco_detalhado[indice_pg] = balanco_detalhado['intercambio'][intercambio_extraido]['programada']
                    ws_balanco_detalhado[indice_vf] = balanco_detalhado['intercambio'][intercambio_extraido]['verificada']                   
                    continue

        wb_ipdo.save('IPDO.xlsx')   # sobrescreve resultados
    
    
    # ENA, EAR e EAM   
    def energial_potencial_armazenada_em_xlsx(self, geral, balanco_detalhado):
        
        print "balanco_detalhado"
        print balanco_detalhado
        wb_ipdo = load_workbook('IPDO.xlsx') 
        ws_balanco_detalhado = wb_ipdo['EnergiaPotencialArmazenada']
        
        ferramenta = Ferramentas()
        
        [primeira_linha, ultima_linha] = ferramenta.linha_nao_vazia(ws_balanco_detalhado)
        
        ultima_linha = ultima_linha + 1        
            
        indice = 'A'+str(ultima_linha)
        
        ws_balanco_detalhado[indice] = geral["data_arquivo"]
    
#        energias_relatorio = ['ENA', 'EAR', 'EAM']
        subsistemas_relatorio = ["sudeste","sul","nordeste","norte"]
        
        
        # referência excel
        coluna_energia_vf = 1 # primeira intercambio verificada
                    
        for subsistema in subsistemas_relatorio:    
            coluna_energia_vf += 1
            coluna_energia_vf
            
            indice_vf = ferramenta.retorna_letra_da_coluna(coluna_energia_vf) + str(ultima_linha)
            print subsistema
            print subsistema
            
            for subsistema_extraido in balanco_detalhado[subsistema]:
                
                if (subsistema_extraido == subsistema):
                    print 'balanco_detalhado'
                    print balanco_detalhado[subsistema_extraido]["eam"]['verificada']
                    ws_balanco_detalhado[indice_vf] = balanco_detalhado[subsistema_extraido]["ena"]['verificada']   
                    ws_balanco_detalhado[indice_vf] = balanco_detalhado[subsistema_extraido]["ear"]['verificada']   
                    ws_balanco_detalhado[indice_vf] = balanco_detalhado[subsistema_extraido]["eam"]['verificada']   
                    continue
                        

#            coluna_energia_vf += 7  # pula fontes programadas
  
        wb_ipdo.save('IPDO.xlsx')   # sobrescreve resultados
        
    
        
        