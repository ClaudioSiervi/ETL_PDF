# -*- coding: utf-8 -*-
"""
Created on Fri Jun 24 11:03:06 2016

@author: Claudio
"""
### ETL --> Extract, Transform and Load


# Convert pdf to text
# http://stackoverflow.com/questions/5725278/how-do-i-use-pdfminer-as-a-library

class Extrair_Transformar_Carregar:
    

    # Desbloqueia um pdf com senha
    # https://lambdafu.net/2011/06/08/remove-password-from-pdf/
    # http://stackoverflow.com/questions/10741600/running-command-lines-within-your-python-script
        
    def desbloqueia(self, entrada, saida):
        # Cria uma versão desbloqueada de PDF com senha
        import os
        os.system('qpdf --decrypt ' + entrada + ' ' + saida)
    
    
    
    def converte_pdf_para_texto(self, path):
    # Converte um arquivo pdf para texto
        from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
        from pdfminer.converter import TextConverter
        from pdfminer.layout import LAParams
        from pdfminer.pdfpage import PDFPage
        from cStringIO import StringIO    
        
        rsrcmgr = PDFResourceManager()
        retstr = StringIO()
        codec = 'utf-8'
        laparams = LAParams()
        device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
        fp = file(path, 'rb')
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        password = ""
        maxpages = 0
        caching = True
        pagenos=set()
    
        for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password,caching=caching, check_extractable=True):
            interpreter.process_page(page)
    
        text = retstr.getvalue()
    
        fp.close()
        device.close()
        retstr.close()
        return text
    


    def converte_pdf_para_html(self, path):
        from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
        from pdfminer.converter import HTMLConverter
        from pdfminer.converter import TextConverter
        from pdfminer.layout import LAParams
        from pdfminer.pdfpage import PDFPage
        from cStringIO import StringIO
        import re
        import csv
        
        
        rsrcmgr = PDFResourceManager()
        retstr = StringIO()
        codec = 'utf-8'
        laparams = LAParams()
        device = HTMLConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
        fp = file(path, 'rb')
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        password = ""
        maxpages = 0 #is for all
        caching = True
        pagenos=set()
        for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password,caching=caching, check_extractable=True):
            interpreter.process_page(page)
        fp.close()
        device.close()
        str = retstr.getvalue()
        retstr.close()
        return str


    def salva_texto_em_txt(self, texto, nome_arquivo_saida):
    # Salva o texto extraído do pdf em um arquivo texto
        self.arquivo_texto = open(nome_arquivo_saida, "w")
        self.status_escrita = self.arquivo_texto.write(texto)
        self.arquivo_texto.close()
        
        
    def salva_texto_em_html(self, texto, nome_arquivo_saida):
    # Salva o texto extraído do pdf em um arquivo texto
        self.arquivo_texto = open(nome_arquivo_saida, "w")
        self.status_escrita = self.arquivo_texto.write(texto)
        self.arquivo_texto.close()
        
        
        
    def salva_texto_em_xlsx(self):
        #from openpyxl import Workbook
        from openpyxl import load_workbook
        
        wb = load_workbook('IPDO.xlsx') 
        ws = wb['Tabela1']
        
        [p, u] = self.linha_nao_vazia(ws)
        
        # overwrite the current document template
        #wb.save('IPDO.xlsx')     
        
        
        
    def linha_nao_vazia(self, ws):
        ######## Encontra a primeira e a ultima linha não vazia
    
        self.tam = ws.max_row + 1
        self.primeira_linha = 0
        self.ultima_linha = 0
        
        for item in xrange(1, self.tam):
            self.celula_valor = ws.cell(row=item, column = 1).value  
            if (self.celula_valor is not None) and (self.primeira_linha == 0):
                self.primeira_linha = item
            elif (self.celula_valor is not None) and (self.primeira_linha is not 0):
                self.ultima_linha = item
        return self.primeira_linha, self.ultima_linha     
        
        
        

        
        
        
         
                
                
        
    
