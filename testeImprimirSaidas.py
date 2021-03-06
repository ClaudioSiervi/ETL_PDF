# -*- coding: utf-8 -*-
"""
Created on Mon Aug 29 12:20:38 2016

@author: Claudio
"""

import os
from bs4 import BeautifulSoup
from Utilitarios import Ferramentas
from openpyxl import load_workbook
from ExtracaoTexto import DadosBalancoEnergeticoDetalhado   
from Mapeamento import DicionarioRegEx

#arquivo_ipdo = ArquivoIPDO(nome_arquivo_entrada)
subsistema = DadosBalancoEnergeticoDetalhado()    
converte = Ferramentas()
caminho = os.getcwd()

caminho = os.getcwd()
caminho = str(caminho) + '\Scripts-py'

for dia in xrange(01,02):
    if (dia <10):
        nome_arquivo_entrada = caminho + '\IPDO-0'+str(dia)+'-05-2016'
    else:
        nome_arquivo_entrada = caminho + '\IPDO-'+str(dia)+'-05-2016'
        
   # exemplo:  nome_arquivo_entrada = 'IPDO-22-06-2016.pdf'
nome_arquivo_saida = nome_arquivo_entrada + '-unlocked.pdf'
# exemplo:  nome_arquivo_entrada = 'IPDO-22-06-2016.pdf'
nome_arquivo_entrada = nome_arquivo_entrada + '.pdf'

tag='div'

html_extraido = converte.pdf_para_html(nome_arquivo_saida)
objeto_bs = BeautifulSoup(html_extraido, 'html.parser')

dic = DicionarioRegEx()
dic = dic.intercambio

inter = subsistema.intercambio_entre_subsistemas(objeto_bs, tag, dic['fontes_lf'], dic['fontes_tp'] )

#print fontes

producao_vf = subsistema.producao(objeto_bs, tag, dic['prod_verif_lf'], dic['prod_verif_tp'] )
producao_pg = subsistema.producao(objeto_bs, tag, dic['prod_prog_lf'], dic['prod_prog_tp'] )


print 'produção_vf  ->' + str(len(producao_vf)) + str(producao_vf)
print 'produção_pg  ->' + str(len(producao_pg)) + str(producao_pg)

if ((len(producao_vf)==5) & (len(producao_pg)==5)):
    carga_vf = [producao_vf.pop()]
    carga_pg = [producao_pg.pop()]

    print 'carga_vf  ->' + str(len(carga_vf)) + '- ' + str(carga_vf)
    print 'carga_pg  ->' + str(len(carga_pg)) + '- ' + str(carga_pg)
   
elif ((len(producao_vf)==4) & (len(producao_pg)==4)):  
    carga_vf = subsistema.producao(objeto_bs, tag, dic['carga_verif_lf'], dic['carga_verif_tp'] )
    carga_pg = subsistema.producao(objeto_bs, tag, dic['carga_prog_lf'], dic['carga_prog_tp'] )
    
    print '5'
    print 'carga_vf  ->' + str(len(carga_vf)) + str(carga_vf)
    print 'carga_pg  ->' + str(len(carga_pg)) + str(carga_pg)
    
else:
    print 'Erro ao ler a carga.'
    print 'O arquivo deve ter mudado de estrutura.'


#ena_vf = subsistema.ena(objeto_bs, tag, dic['ena_lf'], dic['ena_tp'] )
#print 'ena_vf -->' + str(ena_vf)
#
#ear_vf = subsistema.ear(objeto_bs, tag, dic['ear_lf'], dic['ear_tp'] )
#print 'ear_vf -->' + str(ear_vf)


#wb = load_workbook('IPDO.xlsx')
#ws = wb['BalancoEnergeticoDetalhado']
#
#ferramenta = Ferramentas()
#[primeira_linha, ultima_linha] = ferramenta.linha_nao_vazia(ws)
#ultima_linha += 1
##num_elementos_pg = len(producao_vf) # Número de elementos do resumo
##conta_valores = 0
#
#
##letra = ferramenta.retorna_letra_da_coluna(2)
#indice = 'B' + str(ultima_linha)     #Ghidro
#ws[indice] = producao_vf[0]       
#indice = 'C' + str(ultima_linha)     #Gtermo  
#ws[indice] = producao_vf[1] 
#indice = 'D' + str(ultima_linha)     #Gnuclear  
#ws[indice] = producao_vf[2] 
#indice = 'G' + str(ultima_linha)      #Total 
#ws[indice] = producao_vf[3] 
#
#indice = 'H' + str(ultima_linha)      #Total 
#ws[indice] = carga_vf[0] 
#
##
##
##conta_valores = conta_valores + 1
#wb.save('IPDO.xlsx')   # sobrescreve resultados   