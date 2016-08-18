# -*- coding: utf-8 -*-
"""
Created on Thu Jul 28 15:08:38 2016

@author: Claudio
"""

# TESTES
#
#class Data():
#    import datetime
#    mylist = []
#    today = datetime.date.today()
#    mylist.append(today)
#    print mylist[0] # imprime o objeto data, não o container



## Serve apenas para criar novos arquivos.
#import xlsxwriter
#
#workbook = xlsxwriter.Workbook('IPDO.xlsx')
#ws_resultado = workbook.add_worksheet('Tabela1')
#   
from openpyxl import load_workbook  
import string
#        try:
wb = load_workbook('IPDO.xlsx') 
ws = wb['Tabela1']

[plinha, ulinha] = self.linha_nao_vazia(ws)

prevista = string.split(texto[0],';')  
verificada = string.split(texto[1],';')
percent_sin = string.split(texto[2],';')

tam = len(prevista)
conta_valores = 0
for cont in xrange(ulinha,tam):         
    indice = 'B'+ str(cont)            
    ws[indice] = float(prevista[conta_valores])
    conta_valores = conta_valores + 1

wb.save('IPDO.xlsx') 






from bs4 import BeautifulSoup
import string
bs_html = BeautifulSoup(html_extraido, 'html.parser')
t = bs_html.select('div[style*="top:189px"]')
x =t[0].contents
p = ' ' .join(x[0].stripped_strings)
z = p.encode('utf-8')
z = string.split(z, ',')
z = z[1]    # Data do IPDO




######## Encontra a primeira e a ultima linha não vazia
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('IPDO.xlsx') 

ws = wb['Tabela1']

tam = ws.max_row + 1
primeira_linha = 0
ultima_linha = 0

for item in xrange(1, tam):
    celula_valor = ws.cell(row=item, column = 1).value  
    if (celula_valor is not None) and (primeira_linha == 0):
        primeira_linha = item
    elif (celula_valor is not None) and (primeira_linha is not 0):
        ultima_linha = item
        
if (ultima_linha == 0):
    ultima_linha = primeira_linha
    
wb.save('IPDO.xlsx') # sobrescreve resultados













from bs4 import BeautifulSoup


html_doc = """
<html><head><title>The Dormouse's story</title></head>
<body>
<p class="title"><b>The Dormouse's story</b></p>

<p class="story">Once upon a time there were three little sisters; and their names were
<a href="http://example.com/elsie" class="sister" id="link1">Elsie</a>,
<a href="http://example.com/lacie" class="sister" id="link2">Lacie</a> and
<a href="http://example.com/tillie" class="sister" id="link3">Tillie</a>;
and they lived at the bottom of a well.</p>

<p class="story">...</p>
"""


soup = BeautifulSoup(html_extraido, 'html.parser')

print(soup.prettify())


head_tag = soup.head
head_tag
# <head><title>The Dormouse's story</title></head>


title_tag = head_tag.contents[0]
title_tag
# <title>The Dormouse's story</title>
title_tag.contents
# [u'The Dormouse's story']



cont=0
for child in soup.body.descendants:
    cont = cont+1
    soup('div')[cont]

#http://stackoverflow.com/questions/23380171/using-beautifulsoup-extract-text-without-tags
for child in html_extraido.find_all('span'):
    teste = child.text
    if (teste == "*Norte*"):
        print child.text
        
# Imprime somente as tags div com top:505px         
html_extraido.select('div[style*="top:505px"]')        
        
        
        
### Imprime tudo    
for child in html_extraido.find_all('span'):
    print child.text
    
    teste = child.strings
    if (teste == "*Norte*"):
        print 
    
for string in html_extraido.strings:
    print(repr(string))
    
for span in html_extraido.span:
    print(span.string)
    
    
spans = soup.find_all('span')


for each in spans: #iterate over loop [above sections]
    if each.find('spans'):
        continue
    else:
        print each.prettify()
    
    
for row in soup.find_all('div',attrs={"class" : "reviewText"}):
    print row.text
    
        
        
# extrai em um vetor todos os div com  style="top:314px"
t = html_extraido.select('div[style*="top:314px"]')
x =t[0].contents
# imprime somente o texto da primeira tag
resp= ';' .join(x[1].stripped_strings)




^(?=.*\bjack\b)(?=.*\bjames\b).*$



# extrai valores considerando dois atributos da tag div
#http://stackoverflow.com/questions/35140158/using-beautifulsoup-to-find-tag-with-two-specific-styles
import re
y = html_extraido.find('div', style=re.compile(r'left:284px.*?top:314px'))


# Antess
import re
y = html_extraido.find('div', style=re.compile(r'left:284px.*?top:314px'))
x =y.contents
p =''

p = ' ' .join(x[0].stripped_strings)
p = p + ' '
p = p + ' ' .join(x[1].stripped_strings)
p = p + ' '
p = p + ' ' .join(x[2].stripped_strings)
z =p.encode('utf-8')



### Depoissss
left = 'left:' + '284px'
top =  'top:' + '314px'
import re
y = html_extraido.find('div', style=re.compile(r''+ left+'.*?'+top))
x =y.contents
p =''
tam = len(x)
for item in xrange(0, tam):
    p = p + ' ' .join(x[item].stripped_strings)
    p = p + ' '

z = p.encode('utf-8')
print z
    
    
    
    
    
    
import re
import string
from ExtracaoTexto import BalancoEnergia, Subsistemas       
from DicionarioTexto import DicionarioRegEx
from etl_pdf import ExtrairTransformarCarregar
from bs4 import BeautifulSoup
    
objeto_bs = BeautifulSoup(html_extraido, 'html.parser')
tag = 'div'
dados = Subsistemas()    
dic = DicionarioRegEx()

#top_tx =  dic.sudeste['fontes_tp'] 
#left_tx = dic.sudeste['fontes_lf']
top_tx =  dic.sudeste['ear_tp']
left_tx = dic.sudeste['ear_lf'] 
#
#mapeia = ExtrairTransformarCarregar()
#texto_extraido_str = mapeia.extrai_dados_objeto_bs(objeto_bs, tag, left_tx, top_tx)

tag_encontrada = objeto_bs.find(tag, style=re.compile(r''+ left_tx+'.*?'+top_tx))
conteudo_tag = tag_encontrada.contents
texto_extraido_unicode =''
tam = len(conteudo_tag)
for item in xrange(0, tam):
    texto_extraido_unicode = texto_extraido_unicode + ';' .join(conteudo_tag[item].stripped_strings)
    texto_extraido_unicode = texto_extraido_unicode + ';'

texto_extraido_str = texto_extraido_unicode.encode('utf-8')

texto_extraido_str = string.split(texto_extraido_str, ';')
    